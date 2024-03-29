import os
import csv
import openpyxl


def generate_marksheet():

    try:
        os.mkdir('output')
    except:
        pass

    grade_eq = {'AA':10,'AB':9,'BB':8,' BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'DD*':4,'F':0,'F*':0,'I':0,'I*':0}

    # Maping grades...
    with open('grades.csv', 'r') as csvfile:

        reader = csv.DictReader(csvfile)

        grade_record = {}
        for row in reader:
            
            list = []
            list.append(row['Roll'])
            list.append(row['Sem'])
            list.append(row['SubCode'])
            list.append(row['Credit'])
            list.append(row['Grade'])
            list.append(row['Sub_Type'])

            if(grade_record.get(row['Roll']) == None):
                grade_record[row['Roll']] = [list]
            else:
                grade_record[row['Roll']].append(list)

    # Maping names-roll...
    with open('names-roll.csv', 'r') as csvfile:

        reader = csv.DictReader(csvfile)

        name = {}
        for row in reader:
            name[row['Roll']] = row['Name']

    # Maping subjects_master...
    with open('subjects_master.csv', 'r') as csvfile:

        csvreader = csv.DictReader(csvfile)

        subject = {}
        for row in csvreader:
            
            list = []
            list.append(row['subname'])
            list.append(row['ltp'])

            subject[row['subno']] = list


    # Storing result of each student in a nested dictionary...

    result = {}

    for roll in grade_record:
        
        stud_result = {}

        for row in grade_record[roll]:

            sem = row[1]
            subno = row[2]
            credit = row[3]
            grade = row[4]
            sub_type = row[5]

            list = []
            list.append(subno)
            list.append(subject[subno][0])
            list.append(subject[subno][1])
            list.append(credit)
            list.append(sub_type)
            list.append(grade)

            if(stud_result.get(sem) == None):
                stud_result[sem] = [list]
            else:
                stud_result[sem].append(list)
        
        result[roll] = stud_result


    for roll in grade_record:

        # Calculations for Overall sheet...
        semester = []
        spi = []
        cpi = []
        cred_taken = []
        total_cred_taken = []
        total_cred_sum = 0
        cpi_sum = 0

        for i in result[roll]:

            semester.append(i)
            spi_sum = 0
            cred_sum = 0

            for row in result[roll][i]:
                marks = float(grade_eq[row[5]])
                cred = float(row[3])

                spi_sum += marks*cred
                cred_sum += cred
            
            total_cred_sum += cred_sum
            cpi_sum += (spi_sum/cred_sum)*cred_sum
            
            spi.append(round(spi_sum/cred_sum, 2))
            cpi.append(round(cpi_sum/total_cred_sum, 2))
            cred_taken.append(cred_sum)
            total_cred_taken.append(total_cred_sum)

        #  Generating marksheet of each student in output folder... 

        wb = openpyxl.Workbook()

        wb.create_sheet(index = 0, title = 'Overall')
        sheet = wb['Overall']

        sheet.append(['Roll No.', roll])
        sheet.append(['Name of Student', name[roll]])
        sheet.append(['Discipline', roll[4:6]])
        sheet.append(['Semester No.'] + semester)
        sheet.append(['Semester wise Credit Taken'] + cred_taken)
        sheet.append(['SPI'] + spi)
        sheet.append(['Total Credits Taken'] + total_cred_taken)
        sheet.append(['CPI'] + cpi)

        for i in result[roll]:

            t = 'Sem' + i
            wb.create_sheet(index = int(i), title = t)

            sheet = wb[t]

            header = ['Sl No.', 'Subject No.', 'Subject Name', 'L-T-P', 'Credit', 'Subject Type', 'Grade']
            sheet.append(header)

            sno = 0

            for row in result[roll][i]:
                sno += 1
                sheet.append([sno] + row)

        sheetDelete = wb['Sheet']
        wb.remove(sheetDelete)

        path = 'output/' + roll + '.xlsx'

        wb.save(path)

    return


generate_marksheet()